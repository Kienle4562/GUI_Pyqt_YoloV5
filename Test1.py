from openpyxl import load_workbook

wb = load_workbook('Test.xlsm', keep_vba=True, data_only=True)
sh = wb.active
ws = wb['My_sheet']
countA = ws.max_row
print(countA)