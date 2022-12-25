from datetime import date
from datetime import datetime
import serial
import serial.tools.list_ports

from PyQt5 import QtWidgets
from gui import Ui_Form

import math
import shutil
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import threading
import os
import sys
from pathlib import Path
import cv2
import torch
import torch.backends.cudnn as cudnn
import os.path as osp
import numpy as np
import subprocess
import function.helper as helper
import function.utils_rotate as utils_rotate

yolo_LP_detect = torch.hub.load('yolov5', 'custom', path='Model_Yolo\LP_detector.pt', force_reload=True, source='local')
yolo_license_plate = torch.hub.load('yolov5', 'custom', path='Model_Yolo\LP_ocr.pt', force_reload=True, source='local')
yolo_license_plate.conf = 0.60

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from models.common import DetectMultiBackend
from utils.datasets import IMG_FORMATS, VID_FORMATS, LoadImages, LoadStreams
from utils.general import (LOGGER, check_file, check_img_size, check_imshow, check_requirements, colorstr,
                           increment_path, non_max_suppression, print_args, scale_coords, strip_optimizer, xyxy2xywh)
from utils.plots import Annotator, colors, save_one_box
from utils.torch_utils import select_device, time_sync

from deep_sort.utils.parser import get_config  ##K
from deep_sort.deep_sort import DeepSort  ##K
from openpyxl import load_workbook

count = 0  ##K
data = []  ##K


def count_obj(box, w, h, id, classObj, BoxAx , BoxAy, BoxBx, BoxBy):
    global count, data
    today = date.today()
    flag = False
    center_coordinates = (int(box[0]+(box[2]-box[0])/2) , int(box[1]+(box[3]-box[1])/2))
    if center_coordinates[0] > (BoxAx) and center_coordinates[0] < (BoxBx) and center_coordinates[1] > (BoxAy) and center_coordinates[1] < (BoxBy):
        if id not in data:
            flag = True
            count += 1
            data.append(id)
            print('count', count)
            now = datetime.now()
            d1 = today.strftime("%d/%m/%Y")
            current_time = now.strftime("%H:%M:%S")
            wb = load_workbook('Test.xlsm', keep_vba=True)
            # wb = load_workbook('Test.xlsm', keep_vba=True,data_only=True)
            sh = wb.active
            ws = wb['My_sheet']
            countA = ws.max_row
            print('countA',countA)
            countA += 1
            ws.cell(row=countA, column=1).value = d1
            ws.cell(row=countA, column=2).value = current_time
            ws.cell(row=countA, column=3).value = classObj
            wb.save('Test.xlsm')
    return flag


def xyxy2xywh(x):
    # Convert nx4 boxes from [x1, y1, x2, y2] to [x, y, w, h] where xy1=top-left, xy2=bottom-right
    y = x.clone() if isinstance(x, torch.Tensor) else np.copy(x)
    y[:, 0] = (x[:, 0] + x[:, 2]) / 2  # x center
    y[:, 1] = (x[:, 1] + x[:, 3]) / 2  # y center
    y[:, 2] = x[:, 2] - x[:, 0]  # width
    y[:, 3] = x[:, 3] - x[:, 1]  # height
    return y


def xywh2xyxy(x):
    # Convert nx4 boxes from [x, y, w, h] to [x1, y1, x2, y2] where xy1=top-left, xy2=bottom-right
    y = x.clone() if isinstance(x, torch.Tensor) else np.copy(x)
    y[:, 0] = x[:, 0] - x[:, 2] / 2  # top left x
    y[:, 1] = x[:, 1] - x[:, 3] / 2  # top left y
    y[:, 2] = x[:, 0] + x[:, 2] / 2  # bottom right x
    y[:, 3] = x[:, 1] + x[:, 3] / 2  # bottom right y
    return y


def clip_coords(boxes, shape):
    # Clip bounding xyxy bounding boxes to image shape (height, width)
    if isinstance(boxes, torch.Tensor):  # faster individually
        boxes[:, 0].clamp_(0, shape[1])  # x1
        boxes[:, 1].clamp_(0, shape[0])  # y1
        boxes[:, 2].clamp_(0, shape[1])  # x2
        boxes[:, 3].clamp_(0, shape[0])  # y2
    else:  # np.array (faster grouped)
        boxes[:, [0, 2]] = boxes[:, [0, 2]].clip(0, shape[1])  # x1, x2
        boxes[:, [1, 3]] = boxes[:, [1, 3]].clip(0, shape[0])  # y1, y2


def save_one_box_img(xyxy, im, file='image.jpg', gain=1.02, pad=10, square=False, BGR=False, save=True):
    # Save image crop as {file} with crop size multiple {gain} and {pad} pixels. Save and/or return crop
    xyxy = torch.tensor(xyxy).view(-1, 4)
    b = xyxy2xywh(xyxy)  # boxes
    # print("Hello",b)
    if square:
        b[:, 2:] = b[:, 2:].max(1)[0].unsqueeze(1)  # attempt rectangle to square
    b[:, 2:] = b[:, 2:] * gain + pad  # box wh * gain + pad
    xyxy = xywh2xyxy(b).long()
    clip_coords(xyxy, im.shape)
    crop = im[int(xyxy[0, 1]):int(xyxy[0, 3]), int(xyxy[0, 0]):int(xyxy[0, 2]), ::(1 if BGR else -1)]
    # print("file_name", str(increment_path(file).with_suffix('.jpg')))
    if save:
        file.parent.mkdir(parents=True, exist_ok=True)  # make directory
        file_name = str(increment_path(file).with_suffix('.jpg'))
        cv2.imwrite(file_name, crop)

def save_one_box_vid(xyxy, im, file='image.jpg', gain=1.02, pad=10, square=False, BGR=False, save=True, Oj='', LP=False):
    # Save image crop as {file} with crop size multiple {gain} and {pad} pixels. Save and/or return crop
    xyxy = torch.tensor(xyxy).view(-1, 4)
    b = xyxy2xywh(xyxy)  # boxes
    # print("Hello",b)
    if square:
        b[:, 2:] = b[:, 2:].max(1)[0].unsqueeze(1)  # attempt rectangle to square
    b[:, 2:] = b[:, 2:] * gain + pad  # box wh * gain + pad
    xyxy = xywh2xyxy(b).long()
    clip_coords(xyxy, im.shape)
    crop = im[int(xyxy[0, 1]):int(xyxy[0, 3]), int(xyxy[0, 0]):int(xyxy[0, 2]), ::(1 if BGR else -1)]
    print("file_name", str(increment_path(file).with_suffix('.jpg')))
    if LP == True and (Oj == 'car' or Oj == 'Car' or Oj == 'truck' or Oj == 'Truck' or Oj == 'motorcycle'or Oj == "Motorcycle" or Oj == 'bus'or Oj == 'Bus'):
        print("OJ",Oj)
        with open("Sample_OutPut\Image_data_management.txt", "a") as f:
            f.write(str(increment_path(file).with_suffix('.jpg')) + "\n")
    if save:
        file.parent.mkdir(parents=True, exist_ok=True)  # make directory
        file_name = str(increment_path(file).with_suffix('.jpg'))
        cv2.imwrite(file_name, crop)

def detect_lp_img(filename, im0, thickness):
    print("filename", str(filename))
    img = cv2.imread(str(filename))
    plates = yolo_LP_detect(img, size=640)
    list_plates = plates.pandas().xyxy[0].values.tolist()
    list_read_plates = set()
    if len(list_plates) == 0:
        lp = helper.read_plate(yolo_license_plate, img)
        if lp != "unknown":
            cv2.putText(im0, lp, (7, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36,255,12), 2)
            list_read_plates.add(lp)
    else:
        for plate in list_plates:
            flag = 0
            x = int(plate[0])  # xmin
            y = int(plate[1])  # ymin
            w = int(plate[2] - plate[0])  # xmax - xmin
            h = int(plate[3] - plate[1])  # ymax - ymin
            crop_img = im0[y:y + h, x:x + w]
            cv2.rectangle(im0, (int(plate[0]), int(plate[1])),
                          (int(plate[2]), int(plate[3])), color=(0, 0, 225),
                          thickness=thickness)
            # cv2.imwrite("crop.jpg", crop_img)
            # rc_image = cv2.imread("crop.jpg")
            print("dt bsx")
            lp = ""
            for cc in range(0, 2):
                for ct in range(0, 2):
                    lp = helper.read_plate(yolo_license_plate,
                                           utils_rotate.deskew(crop_img, cc, ct))
                    if lp != "unknown":
                        list_read_plates.add(lp)
                        cv2.putText(im0, lp, (int(plate[0]), int(plate[1] - 10)),
                                    cv2.FONT_ITALIC, thickness/3, (36, 255, 12), thickness)
                        flag = 1
                        break
                if flag == 1:
                    break
    cv2.imwrite("images/tmp/single_result.jpg", im0)
    return str(list_read_plates)
    # cv2.imshow('frame', img)
#
# def detect_lp2(xyxy, im, file='image.jpg', gain=1.02, pad=10, square=False, BGR=False, save=True):
#     xyxy = torch.tensor(xyxy).view(-1, 4)
#     b = xyxy2xywh(xyxy)  # boxes
#     # print("Hello",b)
#     if square:
#         b[:, 2:] = b[:, 2:].max(1)[0].unsqueeze(1)  # attempt rectangle to square
#     b[:, 2:] = b[:, 2:] * gain + pad  # box wh * gain + pad
#     xyxy = xywh2xyxy(b).long()
#     clip_coords(xyxy, im.shape)
#     crop = im[int(xyxy[0, 1]):int(xyxy[0, 3]), int(xyxy[0, 0]):int(xyxy[0, 2]), ::(1 if BGR else -1)]
#     print("file_name", str(increment_path(file).with_suffix('.jpg')))
#     if save:
#         file.parent.mkdir(parents=True, exist_ok=True)  # make directory
#         file_name = str(increment_path(file).with_suffix('.jpg'))
#         cv2.imwrite(file_name, crop)
#         # print("Hello2",str(increment_path(file).with_suffix('.jpg')), crop)
#     print("filename", str(file_name))
#     img = cv2.imread(str(file_name))
#     plates = yolo_LP_detect(img, size=640)
#     list_plates = plates.pandas().xyxy[0].values.tolist()
#     list_read_plates = set()
#     if len(list_plates) == 0:
#         lp = helper.read_plate(yolo_license_plate, img)
#         if lp != "unknown":
#             # cv2.putText(img, lp, (7, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36,255,12), 2)
#             list_read_plates.add(lp)
#     else:
#         for plate in list_plates:
#             flag = 0
#             x = int(plate[0])  # xmin
#             y = int(plate[1])  # ymin
#             w = int(plate[2] - plate[0])  # xmax - xmin
#             h = int(plate[3] - plate[1])  # ymax - ymin
#             crop_img = img[y:y + h, x:x + w]
#             # cv2.rectangle(img, (int(plate[0]), int(plate[1])),
#             #               (int(plate[2]), int(plate[3])), color=(0, 0, 225),
#             #               thickness=2)
#             # cv2.imwrite("crop.jpg", crop_img)
#             # rc_image = cv2.imread("crop.jpg")
#             lp = ""
#             for cc in range(0, 2):
#                 for ct in range(0, 2):
#                     lp = helper.read_plate(yolo_license_plate,
#                                            utils_rotate.deskew(crop_img, cc, ct))
#                     if lp != "unknown":
#                         list_read_plates.add(lp)
#                         # cv2.putText(img, lp, (int(plate[0]), int(plate[1] - 10)),
#                         #             cv2.FONT_HERSHEY_SIMPLEX, 0.9, (36, 255, 12), 2)
#                         flag = 1
#                         break
#                 if flag == 1:
#                     break
#     # cv2.imshow('frame', img)
#
#     print(lp)
#     return str(lp)
path = 'images\ImgDrawCoordinates.jpg'
PointA = []
PointB = []
class DrawCoordinates(QWidget):
    def __init__(self):
        super().__init__()
        im0 = cv2.imread(path)
        w, h = im0.shape[1], im0.shape[0]
        self.window_width, self.window_height = w, h
        print("wh",w,h)
        self.setMinimumSize(self.window_width, self.window_height)

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.pix = QPixmap(path)
        # self.pix.fill(Qt.white)
        self.begin, self.destination = QPoint(), QPoint()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawPixmap(QPoint(), self.pix)
        if not self.begin.isNull() and not self.destination.isNull():
            rect = QRect(self.begin, self.destination)
            painter.drawRect(rect.normalized())

    def mousePressEvent(self, event):
        global PointA
        if event.buttons() & Qt.LeftButton:
            print('Point 1')
            self.begin = event.pos()
            self.destination = self.begin
            self.update()
            PointA = str(event.pos())[str(event.pos()).find('(')+1: str(event.pos()).find(')')].split(",")
            print("1", event.pos())

    def mouseMoveEvent(self, event):
        global PointB
        if event.buttons() & Qt.LeftButton:
            print('Point 2')
            self.destination = event.pos()
            self.update()
            PointB = str(event.pos())[str(event.pos()).find('(') + 1: str(event.pos()).find(')')].split(",")
            print("2", event.pos())

class MainWindows(QtWidgets.QWidget, Ui_Form):
    def __init__(self):
        super(MainWindows, self).__init__()
        self.windowDrawCoordinates = DrawCoordinates()
        self.setupUi(self)
        self.init()
        self.setWindowTitle("Yolo Host Computer")
        self.ser = serial.Serial()
        # self.port_check()
        self.setWindowIcon(QIcon("images/UI/logo.jpg"))

        # The number of received data and sent data is set to zero
        # self.data_num_received = 0
        # self.data_num_sended = 0
        # self.rec_lcdNumber.display(self.data_num_received)
        # self.send_lcdNumber.display(self.data_num_sended)

        # Image reading process
        self.output_size = 480
        self.img2predict = ""
        self.device = 'cpu'
        self.deviceName = 'CPU'
        self.classes = [0]

        self.pointAx = '0'
        self.pointAy = '200'
        self.pointBx = '640'
        self.pointBy = '200'
        self.checkGPU = str(torch.cuda.is_available())
        self.model_path = "Model_Yolo/yolov5n.pt"
        # Initialize the video read thread
        self.vid_source = '0'  # The initial setting is the camera
        self.stopEvent = threading.Event()
        self.webcam = True
        self.stopEvent.clear()

        # self.model = self.model_load(weights=self.model_path,
        #                              device=self.device)
        # todo A device indicating where the model is loaded
        self.model = self.model_load(weights="Model_Yolo/yolov5n.pt",
                                     device='cpu')

        self.reset_vid()
        self.line_thickness = 3
        self.status_hide_labels = False
        self.status_hide_conf = False
        self.status_save_crop = False
        self.status_license_plate_recogniton = False

        self.pushButton_streaming.setEnabled(False)
        self.pushButton_loadmp4.setEnabled(False)
        self.pushButton_stopscan.setEnabled(False)
        self.pushButton_loadpic.setEnabled(False)
        self.pushButton_scanpic.setEnabled(False)
        self.pushButton_comparePic.setEnabled(False)

    '''
    ***Model initialization***
    '''
    def showObjCodesTable(self, checked):
        self.label_video.setPixmap(QPixmap("images/ObjCodesTable.jpg"))
        self.label_video.setScaledContents(True)
        self.left_img.setPixmap(QPixmap("images/ObjCodesTable.jpg"))
        self.left_img.setScaledContents(True)

    def toggle_windowDrawCoordinates(self, checked):
        if self.windowDrawCoordinates.isVisible():
            self.windowDrawCoordinates.hide()

        else:
            self.windowDrawCoordinates.show()

    # Select image config file to read
    def importPicSampleCoordinates(self, checked):
        fileName, fileType = QFileDialog.getOpenFileName(self, 'Choose file', '', '*.jpg *.png *.tif *.jpeg')
        if fileName:
            shutil.copy(fileName, "images/ImgDrawCoordinates.jpg")
            print("File copied successfully:",fileName)
            self.windowDrawCoordinates = DrawCoordinates()

    @torch.no_grad()
    def model_load(self, weights="",  # model.pt path(s)
                   device='',  # cuda device, i.e. 0 or 0,1,2,3 or cpu
                   half=False,  # use FP16 half-precision inference
                   dnn=False,  # use OpenCV DNN for ONNX inference
                   ):
        device = select_device(device)
        half &= device.type != 'cpu'  # half precision only supported on CUDA
        device = select_device(device)
        model = DetectMultiBackend(weights, device=device, dnn=dnn)
        stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
        # Half
        half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
        if pt:
            model.model.half() if half else model.model.float()
            print("Mode!", weights)
        print("Model loading is complete!", weights)

        return model

    def reset_vid(self):
        self.pushButton_streaming.setEnabled(True)
        self.pushButton_loadmp4.setEnabled(True)
        self.label_video.setPixmap(QPixmap("images/UI/yolo.jpg"))

        self.label_video.setScaledContents(True)
        # self.vid_source = '0'
        self.webcam = True

    def init(self):
        # Serial port detection button
        # self.box_1.clicked.connect(self.port_check)
        # pf.test(100,0.3)
        # Serial port information display
        # self.box_2.currentTextChanged.connect(self.port_imf)

        # Open serial button
        # self.open_button.clicked.connect(self.port_open)

        # Close serial button
        # self.close_button.clicked.connect(self.port_close)

        # Send data button
        # self.send_button.clicked.connect(self.data_send)

        # Send data regularly
        # self.timer_send = QTimer()
        # self.timer_send.timeout.connect(self.data_send)
        # self.timer_send_cb.stateChanged.connect(self.data_send_timer)

        # Timer receives data
        # self.timer = QTimer(self)
        # self.timer.timeout.connect(self.data_receive)

        # Clear send window
        # self.clear_button.clicked.connect(self.send_data_clear)

        # Clear receive window
        # self.clear_button.clicked.connect(self.receive_data_clear)

        # Upload image window
        self.pushButton_loadpic.clicked.connect(self.upload_img)

        # Detect picture window
        self.pushButton_scanpic.clicked.connect(self.detect_img)
        # Detect video window
        self.pushButton_streaming.clicked.connect(self.open_cam)
        self.pushButton_loadmp4.clicked.connect(self.open_mp4)
        self.pushButton_stopscan.clicked.connect(self.close_vid)

        self.pushButton_upload_yolo.clicked.connect(self.open_model_yolo)
        self.buttonUpdateSettings.clicked.connect(self.yolo_configuration_settings)
        self.buttonUploadFilePicSample.clicked.connect(self.importPicSampleCoordinates)
        self.buttonDrawCoordinates.clicked.connect(self.toggle_windowDrawCoordinates)
        self.buttonShowObjCodesTable.clicked.connect(self.showObjCodesTable)

    '''
    ***Upload image***
    '''
    def yolo_configuration_settings(self):
        # self.checkBox_Settings.setEnabled(False)
        global PointA, PointB
        number1 = self.select_input.text()
        numberAx = self.point_Ax.text()
        numberAy = self.point_Ay.text()
        numberBx = self.point_Bx.text()
        numberBy = self.point_By.text()
        number4 = self.box_thickness.text()
        flag_err = False

        # SelectCamera
        if self.select_input.text() != "":
            try:
                number1 = int(number1)
                cap = cv2.VideoCapture(number1, cv2.CAP_DSHOW)
                if (cap.isOpened()):
                    self.vid_source = str(self.select_input.text())
                else:
                    QMessageBox.about(self, 'Warning', "Can not detect the camera")
                    flag_err = True
            except Exception:
                QMessageBox.about(self, 'Error', 'Input can only be a number \nEx:0 or 1 or...')
                flag_err = True
                pass
        else:
            self.vid_source = '0'
        self.select_input.setText(str(self.vid_source))

        # Select GPU
        if str(self.select_GPU.currentText()) == 'Graphics Card':
            if self.checkGPU:
                self.device = str('0')  # GPU
                self.deviceName = str('Graphics Card')
            else:
                QMessageBox.about(self, 'Warning', 'GPU not found. Please use CPU')
                flag_err = True
        else:
            self.device = str('cpu')  # CPU
            self.deviceName = str('CPU')

        # Select Object
        list = self.choose_object.text().split(",")
        list2 = []
        for i in range(len(list)):
            if list[i] != "" and list[i].isnumeric():
                t = int(list[i])
                list2.append(t)
        if list2 == []: list2 = [0]
        self.classes = list2
        self.choose_object.setText(str(self.classes).replace("[","").replace(" ","").replace("]",""))
        # Box Thickness
        if self.box_thickness.text() != "":
            try:
                number4 = int(number4)
                self.line_thickness = int(self.box_thickness.text())
            except Exception:
                QMessageBox.about(self, 'Error', 'Input box thickness can only be a number')
                flag_err = True
                pass
        else:
            self.line_thickness = 1
        self.box_thickness.setText(str(self.line_thickness))

        # Point Ax
        if self.point_Ax.text() != "":
            try:
                numberAx = int(numberAx)
                self.pointAx = int(self.point_Ax.text())
                # if len(PointA) < 1: PointA[0] = int(self.point_Ax.text())
            except Exception:
                QMessageBox.about(self, 'Error', 'Input point left(Ax) can only be number')
                flag_err = True
                pass
        else:
            self.pointAx = 115
            # if len(PointA) < 1: PointA[0] = 0
        self.point_Ax.setText(str(self.pointAx))

        # Point Ay
        if self.point_Ay.text() != "":
            try:
                numberAy = int(numberAy)
                self.pointAy = int(self.point_Ay.text())
                # if len(PointA) < 2:PointA[1] = int(self.point_Ay.text())
            except Exception:
                QMessageBox.about(self, 'Error', 'Input point left(Ax) can only be number')
                flag_err = True
                pass
        else:
            self.pointAy = 80
            # if len(PointA) < 2:PointA[1] = 250
        self.point_Ay.setText(str(self.pointAy))

        # Point Bx
        if self.point_Bx.text() != "":
            try:
                numberBx = int(numberBx)
                self.pointBx = int(self.point_Bx.text())
                # if len(PointB) < 1: PointB[0] = int(self.point_Bx.text())
            except Exception:
                QMessageBox.about(self, 'Error', 'Input point right(Bx) can only be a number')
                flag_err = True
                pass
        else:
            self.pointBx = 525
            # if len(PointB) < 1: PointB[0] = 640
        self.point_Bx.setText(str(self.pointBx))
        # Point By
        if self.point_By.text() != "":
            try:
                numberBy = int(numberBy)
                self.pointBy = int(self.point_By.text())
                # if len(PointB) < 2: PointB[1] = int(self.point_By.text())
            except Exception:
                QMessageBox.about(self, 'Error', 'Input point right(Bx) can only be a number')
                flag_err = True
                pass
        else:
            self.pointBy = 635
            # if len(PointB) < 2: PointB[1] = 250
        self.point_By.setText(str(self.pointBy))

        # Show pos
        if PointA and PointB:
            print("PointAB",PointA,"\n", PointB)
            self.point_Ax.setText(str(PointA[0]))
            self.pointAx = int(PointA[0])
            self.point_Ay.setText(str(PointA[1]))
            self.pointAy = int(PointA[1])
            self.point_Bx.setText(str(PointB[0]))
            self.pointBx = int(PointB[0])
            self.point_By.setText(str(PointB[1]))
            self.pointBy = int(PointB[1])

        # Load model #Select Yolo Model
        basename = os.path.basename(self.model_path)
        self.pushButton_upload_yolo.setText(str(basename))
        fileName = self.model_path
        if fileName != "":
            self.model = self.model_load(weights=str(fileName),
                                         device=str(self.device)) #
            print("Upload model yolo complete:", str(fileName))
            self.pushButton_upload_yolo.setText(basename)

        # Hide Labels
        if self.hide_labels.checkState() > 0:
            self.status_hide_labels = True
        else:
            self.status_hide_labels = False

        # Hide Confidences
        if self.hide_confidences.checkState() > 0:
            self.status_hide_conf = True
        else:
            self.status_hide_conf = False

        # License Plate Recognition
        if self.license_plate_recogniton.checkState() > 0:
            self.status_license_plate_recogniton = True
            # with open("Sample_OutPut\Image_data_management.txt", "w") as f:
            #     f.write("LP_On" + "\n")
        else:
            self.status_license_plate_recogniton = False
            # with open("Sample_OutPut\Image_data_management.txt", "w") as f:
            #     f.write("LP_Off" + "\n")

        # Save Image Data
        if self.save_img_data.checkState() > 0:
            self.status_save_crop = True
        else:
            self.status_save_crop = False

        # Show result
        text_br = "Cam:" + str(self.vid_source) + " | Yolo Model:" + str(basename) + " | GPU:" + str(
            self.deviceName) + " | Object:" + str(self.classes) + '\n' + "Point Right:" + str(
            self.pointBy) + " | Point Left:" + str(self.pointAy) + " | Hide Labels:" + str(
            self.status_hide_labels) + " | Hide Confidences:" + str(
            self.status_hide_conf) + '\n' + "License Plate Recognition:" + str(
            self.status_license_plate_recogniton) + " | Save Image Data:" + str(self.status_save_crop)
        self.textBrowser_video.setText(str(text_br))
        self.textBrowser_pic.setText(str(text_br))

        if not flag_err:
            QMessageBox.about(self, 'Complete', 'Configuration has been updated')

            self.pushButton_streaming.setEnabled(True)
            self.pushButton_loadmp4.setEnabled(True)
            self.pushButton_stopscan.setEnabled(True)
            self.pushButton_loadpic.setEnabled(True)
            self.pushButton_scanpic.setEnabled(True)
            self.pushButton_comparePic.setEnabled(True)


    def open_model_yolo(self):
        fileName, fileType = QFileDialog.getOpenFileName(self, 'Choose file', '', '*.pt')
        filepath_not_exist = str(fileName)
        basename = os.path.basename(filepath_not_exist)  # basename = name ex: yolov5.pt
        if fileName != "":
            self.model_path = fileName
            print("Upload model yolo complete:", str(fileName))
            self.textBrowser_pic.setText("Upload model yolo complete")
            self.textBrowser_video.setText("Upload model yolo complete")
            self.pushButton_upload_yolo.setText(basename)

    def upload_img(self):
        # Select the video file to read
        fileName, fileType = QFileDialog.getOpenFileName(self, 'Choose file', '', '*.jpg *.png *.tif *.jpeg')
        if fileName:
            suffix = fileName.split(".")[-1]
            save_path = osp.join("images/tmp", "tmp_upload." + suffix)
            shutil.copy(fileName, save_path)
            # You should resize the pictures and put them together
            im0 = cv2.imread(save_path)
            resize_scale = self.output_size / im0.shape[0]
            im0 = cv2.resize(im0, (0, 0), fx=resize_scale, fy=resize_scale)
            cv2.imwrite("images/tmp/upload_show_result.jpg", im0)
            # self.right_img.setPixmap(QPixmap("images/tmp/single_result.jpg"))
            self.img2predict = fileName
            self.left_img.setPixmap(QPixmap("images/tmp/upload_show_result.jpg"))
            self.left_img.setScaledContents(True)

            # todo The image on the right is reset after uploading the image，
            self.right_img.setPixmap(QPixmap("images/UI/logo.jpg"))
            self.right_img.setScaledContents(True)

    '''
    ***Detect pictures***
    '''

    def detect_img(self):
        self.select_input.setText(str(self.vid_source))
        self.point_Ax.setText('0')
        self.point_Ay.setText('0')
        self.point_Bx.setText('0')
        self.point_By.setText('0')
        self.pushButton_upload_yolo.setText(str(os.path.basename(self.model_path)))

        model = self.model
        output_size = self.output_size
        source = self.img2predict  # file/dir/URL/glob, 0 for webcam
        imgsz = 640  # inference size (pixels)
        conf_thres = 0.25  # confidence threshold
        iou_thres = 0.45  # NMS IOU threshold
        max_det = 1000  # maximum detections per image
        # device = self.device  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        view_img = False  # show results
        save_txt = False  # save results to *.txt
        save_conf = False  # save confidences in --save-txt labels
        save_crop = self.status_save_crop  # save cropped prediction boxes
        nosave = False  # do not save images/videos
        classes = self.classes  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms = False  # class-agnostic NMS
        augment = False  # ugmented inference
        visualize = False  # visualize features
        line_thickness = self.line_thickness  # bounding bo thickness (pixels)
        hide_labels = self.status_hide_labels  # hide labels
        hide_conf = self.status_hide_conf  # hide confidences
        half = False  # use FP16 half-precision inference
        dnn = False  # use OpenCV DNN for ONNX inference
        print(source)
        if source == "":
            QMessageBox.warning(self, "Please Upload", "Please Upload Pictures Before Testing")
        else:
            source = str(source)
            device = select_device(self.device)
            webcam = False
            stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
            imgsz = check_img_size(imgsz, s=stride)  # check image size
            save_img = not nosave and not source.endswith('.txt')  # save inference images
            # Dataloader
            if webcam:
                view_img = check_imshow()
                cudnn.benchmark = True  # set True to speed up constant image size inference
                dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt and not jit)
                bs = len(dataset)  # batch_size
            else:
                dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
                bs = 1  # batch_size
            vid_path, vid_writer = [None] * bs, [None] * bs
            # Run inference
            if pt and device.type != 'cpu':
                print("Issues warmup")
                # model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
            dt, seen = [0.0, 0.0, 0.0], 0
            for path, im, im0s, vid_cap, s in dataset:
                t1 = time_sync()
                im = torch.from_numpy(im).to(device)
                im = im.half() if half else im.float()  # uint8 to fp16/32
                im /= 255  # 0 - 255 to 0.0 - 1.0
                if len(im.shape) == 3:
                    im = im[None]  # expand for batch dim
                t2 = time_sync()
                dt[0] += t2 - t1
                # Inference
                # visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
                pred = model(im, augment=augment, visualize=visualize)
                t3 = time_sync()
                dt[1] += t3 - t2
                # NMS
                pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
                dt[2] += time_sync() - t3
                # Second-stage classifier (optional)
                # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)
                # Process predictions
                self.textBrowser_pic.setText("")
                for i, det in enumerate(pred):  # per image
                    seen += 1
                    if webcam:  # batch_size >= 1
                        p, im0, frame = path[i], im0s[i].copy(), dataset.count
                        s += f'{i}: '
                    else:
                        p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)
                    p = Path(p)  # to Path
                    s += '%gx%g ' % im.shape[2:]  # print string
                    gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
                    imc = im0.copy() if save_crop else im0  # for save_crop
                    annotator = Annotator(im0, line_width=line_thickness, example=str(names))
                    count_object = 0
                    if len(det):
                        # Rescale boxes from img_size to im0 size
                        det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()
                        # Print results
                        for c in det[:, -1].unique():
                            n = (det[:, -1] == c).sum()  # detections per class
                            s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string
                        # Write results
                        for *xyxy, conf, cls in reversed(det):
                            count_object = count_object + 1
                            if save_txt:  # Write to file
                                xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(
                                    -1).tolist()  # normalized xywh
                                line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                                # with open(txt_path + '.txt', 'a') as f:
                                #     f.write(('%g ' * len(line)).rstrip() % line + '\n')
                            if save_img or save_crop or view_img:  # Add bbox to image
                                c = int(cls)  # integer class
                                label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                                str_box = annotator.box_label(xyxy, label, color=colors(c, True))

                                # t = str(xyxy)+str(label)
                                t = str(label) + " | Object has been counted: " + str(count_object)
                                with open("temp.txt", "w") as f:
                                    f.write(t + "\n")
                                if not self.status_license_plate_recogniton: self.textBrowser_pic.setText(t)
                                save_dir = increment_path(Path("Sample_OutPut") / 'exp',
                                                          exist_ok=True)  # increment run
                                save_dir.mkdir(parents=True, exist_ok=True)  # make dir
                                if save_crop:
                                    save_one_box_img(xyxy, imc, file=save_dir / 'crops' / names[
                                        c] / datetime.now().strftime(
                                        "d" + "%d%m%Y") / f'{datetime.now().strftime("%H%M%S") + "_"}.jpg',
                                                     BGR=True)

                    LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')
                    # Stream results
                    imlp = annotator.result()
                    im0 = annotator.result()
                    # if view_img:
                    #     cv2.imshow(str(p), im0)
                    #     cv2.waitKy(1)  # 1 millisecond
                    # Save results (image with detections)
                    resize_scale = output_size / im0.shape[0]
                    im0 = cv2.resize(im0, (0, 0), fx=resize_scale, fy=resize_scale)
                    # cv2.imwrite("images/tmp/single_result.jpg", im0)
                    if count_object <=0: self.textBrowser_pic.setText('Unable to recognize object')
                    if self.status_license_plate_recogniton and count_object > 1: # Run 1 time
                        if names[c] == "car" or "motorcycle":
                            self.textBrowser_pic.setText(str(open("temp.txt", "r").read()) + detect_lp_img(filename='images/tmp/tmp_upload.jpg', im0 = imlp, thickness=line_thickness))
                        else:
                            cv2.imwrite("images/tmp/single_result.jpg", im0)
                    else:
                        cv2.imwrite("images/tmp/single_result.jpg", im0)

                    # From the current situation, it should only be a problem under ubuntu, but it is complete under windows, so continue
                    # if self.checkBox_circle.checkState() > 0:
                    #     # read input
                    #     img = cv2.imread('images/tmp/upload_show_result.jpg')
                    #     # convert to gray
                    #     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    #     # threshold
                    #     thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)[1]
                    #     # find largest contour
                    #     contours = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    #     contours = contours[0] if len(contours) == 2 else contours[1]
                    #     if len(contours) != 0:
                    #         big_contour = max(contours, key=cv2.contourArea)
                    #         # fit contour to ellipse and get ellipse center, minor and major diameters and angle in degree
                    #         ellipse = cv2.fitEllipse(big_contour)  # big_contour)
                    #         (xc, yc), (d1, d2), angle = ellipse
                    #         # print(xc, yc, d1, d1, angle)
                    #         # print("big_contour",big_contour)
                    #         # draw ellipse
                    #         result = img.copy()
                    #         cv2.ellipse(result, ellipse, (0, 255, 0), 3)
                    #         # draw circle at center
                    #         xc, yc = ellipse[0]
                    #         cv2.circle(result, (int(xc), int(yc)), 10, (255, 255, 255), -1)
                    #         # draw vertical line
                    #         # compute major radius
                    #         rmajor = max(d1, d2) / 2
                    #         if angle > 90:
                    #             angle = angle - 90
                    #         else:
                    #             angle = angle + 90
                    #         print(angle)
                    #         xtop = xc + math.cos(math.radians(angle)) * rmajor
                    #         ytop = yc + math.sin(math.radians(angle)) * rmajor
                    #         xbot = xc + math.cos(math.radians(angle + 180)) * rmajor
                    #         ybot = yc + math.sin(math.radians(angle + 180)) * rmajor
                    #         cv2.line(result, (int(xtop), int(ytop)), (int(xbot), int(ybot)), (0, 0, 255), 3)
                    #         cv2.imwrite("images/tmp/single_result.jpg", result)
                    #     else:
                    #         cv2.imwrite("images/tmp/single_result.jpg", thresh)
                    self.right_img.setPixmap(QPixmap("images/tmp/single_result.jpg"))
                    self.right_img.setScaledContents(True)

    # Video detection, the logic is basically the same, there are two functions, namely, the function of detecting the camera and the function of detecting video files, and the function of detecting the camera first。

    '''
    ### UI close event ###
    '''

    def closeEvent(self, event):
        reply = QMessageBox.question(self,
                                     'Quit',
                                     "Are you sure?",
                                     QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.close()
            event.accept()
        else:
            event.ignore()

    '''
    ### Video off event ###
    '''

    def open_cam(self):
        self.pushButton_streaming.setEnabled(False)
        self.pushButton_loadmp4.setEnabled(False)
        self.pushButton_stopscan.setEnabled(True)
        # self.vid_source = '0'
        self.webcam = True
        # Reset the button to him
        # print("GOGOGO")
        th = threading.Thread(target=self.detect_vid)
        th.start()

    '''
    ### Enable video file detection event ###
    '''

    def open_mp4(self):
        fileName, fileType = QFileDialog.getOpenFileName(self, 'Choose file', '', '*.mp4 *.avi')
        if fileName:
            self.pushButton_streaming.setEnabled(False)
            self.pushButton_loadmp4.setEnabled(False)
            # self.pushButton_stopscan.setEnabled(True)
            self.vid_source = fileName
            self.webcam = False
            th = threading.Thread(target=self.detect_vid)
            th.start()

    '''
    ### Video start event ###
    '''

    # The main functions of the video and the camera are the same, but the incoming source is different.

    def detect_vid(self):
        self.buttonUpdateSettings.setEnabled(False)
        # self.lineEdit.setText(str(self.vid_source)) #CAM
        self.pushButton_upload_yolo.setText(str(os.path.basename(self.model_path)))  # Model Yolo
        self.point_Ay.setText(str(self.pointAy))  # Point Right
        self.point_By.setText(str(self.pointBy))  # Point Left
        # if self.status_license_plate_recogniton: subprocess.Popen('D:\GUI_Pyqt5_YoloV5\RunLP.bat',
        #                                                           creationflags=subprocess.CREATE_NEW_CONSOLE)
        model = self.model
        print("MD", self.model_path)
        output_size = self.output_size
        # source = self.img2predict  # file/dir/URL/glob, 0 for webcam
        imgsz = 640  # inference size (pixels)
        conf_thres = 0.25  # confidence threshold
        iou_thres = 0.45  # NMS IOU threshold
        max_det = 1000  # maximum detections per image
        # device = self.device  # cuda device, i.e. 0 or 0,1,2,3 or cpu
        view_img = False  # show results
        save_txt = False  # save results to *.txt
        save_conf = False  # save confidences in --save-txt labels
        save_crop = self.status_save_crop  # save cropped prediction boxes
        nosave = False  # do not save images/videos
        classes = self.classes  # None  # filter by class: --class 0, or --class 0 2 3
        agnostic_nms = False  # class-agnostic NMS
        augment = False  # ugmented inference
        visualize = False  # visualize features
        line_thickness = int(self.line_thickness)  # bounding box thickness (pixels)
        hide_labels = self.status_hide_labels  # hide labels
        hide_conf = self.status_hide_conf  # hide confidences
        half = False  # use FP16 half-precision inference
        dnn = False  # use OpenCV DNN for ONNX inference
        source = str(self.vid_source)
        webcam = self.webcam
        device = select_device(self.device)
        stride, names, pt, jit, onnx = model.stride, model.names, model.pt, model.jit, model.onnx
        imgsz = check_img_size(imgsz, s=stride)  # check image size
        save_img = not nosave and not source.endswith('.txt')  # save inference images

        config_deepsort = 'deep_sort/configs/deep_sort.yaml'  ##K
        deep_sort_model = 'osnet_x0_25'  ##

        project = "Sample_OutPut"
        name = 'exp'
        exist_ok = False
        save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
        save_dir.mkdir(parents=True, exist_ok=True)  # make dir

        txt_file_name = source.split('/')[-1].split('.')[0]
        txt_path = str(Path(save_dir)) + '/' + txt_file_name + '.txt'

        # Dataloader
        if webcam:
            view_img = check_imshow()
            cudnn.benchmark = True  # set True to speed up constant image size inference
            dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt and not jit)
            bs = len(dataset)  # batch_size
        else:
            dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt and not jit)
            bs = 1  # batch_size
        vid_path, vid_writer = [None] * bs, [None] * bs
        # Run inference
        cfg = get_config()  ##K
        cfg.merge_from_file(config_deepsort)
        deepsort = DeepSort(deep_sort_model,
                            max_dist=cfg.DEEPSORT.MAX_DIST,
                            max_iou_distance=cfg.DEEPSORT.MAX_IOU_DISTANCE,
                            max_age=cfg.DEEPSORT.MAX_AGE, n_init=cfg.DEEPSORT.N_INIT, nn_budget=cfg.DEEPSORT.NN_BUDGET,
                            use_cuda=True)  ##K

        device = select_device(device)
        half &= device.type != 'cpu'

        half &= pt and device.type != 'cpu'  # half precision only supported by PyTorch on CUDA
        if pt:
            model.model.half() if half else model.model.float()

        if pt and device.type != 'cpu':
            print("Issues warmup")
            # model(torch.zeros(1, 3, *imgsz).to(device).type_as(next(model.model.parameters())))  # warmup
        dt, seen = [0.0, 0.0, 0.0, 0.0], 0
        for path, im, im0s, vid_cap, s in dataset:
            t1 = time_sync()
            im = torch.from_numpy(im).to(device)
            im = im.half() if half else im.float()  # uint8 to fp16/32
            im /= 255  # 0 - 255 to 0.0 - 1.0
            if len(im.shape) == 3:
                im = im[None]  # expand for batch dim
            t2 = time_sync()
            dt[0] += t2 - t1
            # Inference
            # visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
            pred = model(im, augment=augment, visualize=visualize)
            t3 = time_sync()
            dt[1] += t3 - t2
            # NMS
            pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)
            dt[2] += time_sync() - t3
            # Second-stage classifier (optional)
            # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)
            # Process predictions
            for i, det in enumerate(pred):  # per image
                seen += 1
                if webcam:  # batch_size >= 1
                    p, im0, frame = path[i], im0s[i].copy(), dataset.count
                    s += f'{i}: '
                else:
                    p, im0, frame = path, im0s.copy(), getattr(dataset, 'frame', 0)
                p = Path(p)  # to Path
                # save_path = str(save_dir / p.name)  # im.jpg
                # txt_path = str(save_dir / 'labels' / p.stem) + (
                #     '' if dataset.mode == 'image' else f'_{frame}')  # im.txt
                s += '%gx%g ' % im.shape[2:]  # print string
                gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
                imc = im0.copy() if save_crop or self.status_license_plate_recogniton else im0  # for save_crop

                cv2.imwrite("images/tmp/single_org_vid.jpg", im0)
                annotator = Annotator(im0, line_width=line_thickness, example=str(names))
                w, h = im0.shape[1], im0.shape[0]  ##Kx
                if det is not None and len(det):
                    # Rescale boxes from img_size to im0 size
                    det[:, :4] = scale_coords(im.shape[2:], det[:, :4], im0.shape).round()

                    # Print results
                    for c in det[:, -1].unique():
                        n = (det[:, -1] == c).sum()  # detections per class
                        s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string
                    xywhs = xyxy2xywh(det[:, 0:4])  ##K
                    confs = det[:, 4]  ##K
                    clss = det[:, 5]  ##K

                    t4 = time_sync()  ##K
                    outputs = deepsort.update(xywhs.cpu(), confs.cpu(), clss.cpu(), im0)  ##K
                    t5 = time_sync()  ##K
                    dt[3] += t5 - t4  ##K

                    # Implement only det_max with the highest confidence in all annotation boxes
                    MaxConf = []
                    len_det = len(det)

                    if len(outputs) > 0:
                        for j, (output, conf) in enumerate(zip(outputs, confs)):
                            bboxes = output[0:4]
                            id = output[4]
                            cls = output[5]
                            # count
                            if count_obj(bboxes, w, h, id, classObj = int(cls), BoxAx=self.pointAx, BoxAy=self.pointAy, BoxBx=self.pointBx, BoxBy=self.pointBy) and self.status_license_plate_recogniton:
                            # count_obj(bboxes, w, h, id)
                            # if self.status_license_plate_recogniton:
                                if os.stat("dataset_output\List_Result_License_Plate_Recognition.txt").st_size > 0:
                                    with open("dataset_output\List_Result_License_Plate_Recognition.txt", 'r') as f:
                                        for countLP, lineLP in enumerate(f):
                                            pass
                                    lineLP = countLP + 1
                                    print('line', lineLP)
                                    f = open("dataset_output\List_Result_License_Plate_Recognition.txt")
                                    list_contentLP = f.readlines()
                                    if lineLP >=1:
                                        self.label_lp_9.setPixmap(QPixmap(str(list_contentLP[lineLP - 1]).replace("\n","")))
                                        self.label_lp_9.setScaledContents(True)
                                        textLP9 = list_contentLP[lineLP - 1].split('/')
                                        textLP9 = textLP9[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_9.setText(str('textLP9'))
                                    if lineLP >= 2:
                                        self.label_lp_8.setPixmap(QPixmap(str(list_contentLP[lineLP - 2].replace("\n",""))))
                                        self.label_lp_8.setScaledContents(True)
                                        textLP8 = list_contentLP[lineLP - 2].split('/')
                                        textLP8 = textLP8[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_8.setText(str('textLP8'))

                                    if lineLP >= 3:
                                        self.label_lp_7.setPixmap(QPixmap(str(list_contentLP[lineLP - 3].replace("\n",""))))
                                        self.label_lp_7.setScaledContents(True)
                                        textLP7 = list_contentLP[lineLP - 3].split('/')
                                        textLP7 = textLP7[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_7.setText(str('textLP7'))
                                    if lineLP >= 4:
                                        self.label_lp_6.setPixmap(QPixmap(str(list_contentLP[lineLP - 4].replace("\n",""))))
                                        self.label_lp_6.setScaledContents(True)
                                        textLP6 = list_contentLP[lineLP - 4].split('/')
                                        textLP6 = textLP6[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_6.setText(str(textLP6))
                                    if lineLP >= 5:
                                        self.label_lp_5.setPixmap(QPixmap(str(list_contentLP[lineLP - 5].replace("\n",""))))
                                        self.label_lp_5.setScaledContents(True)
                                        textLP5 = list_contentLP[lineLP - 5].split('/')
                                        textLP5 = textLP5[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_5.setText(str(textLP5))
                                    if lineLP >= 6:
                                        self.label_lp_4.setPixmap(QPixmap(str(list_contentLP[lineLP - 6].replace("\n",""))))
                                        self.label_lp_4.setScaledContents(True)
                                        textLP4 = list_contentLP[lineLP - 6].split('/')
                                        textLP4 = textLP4[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_4.setText(str(textLP4))
                                    if lineLP >= 7:
                                        self.label_lp_3.setPixmap(QPixmap(str(list_contentLP[lineLP - 7].replace("\n",""))))
                                        self.label_lp_3.setScaledContents(True)
                                        textLP3 = list_contentLP[lineLP - 7].split('/')
                                        textLP3 = textLP3[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_3.setText(str(textLP3))
                                    if lineLP >= 8:
                                        self.label_lp_2.setPixmap(QPixmap(str(list_contentLP[lineLP - 8].replace("\n",""))))
                                        self.label_lp_2.setScaledContents(True)
                                        textLP2 = list_contentLP[lineLP - 8].split('/')
                                        textLP2 = textLP2[2].replace('.jpg','').replace("\n","")

                                        # self.text_label_lp_2.setText(str(textLP2))
                                    if lineLP >= 9:
                                        self.label_lp_1.setPixmap(QPixmap(str(list_contentLP[lineLP - 9].replace("\n",""))))
                                        self.label_lp_1.setScaledContents(True)
                                        textLP1 = list_contentLP[lineLP - 9].split('/')
                                        textLP1 = textLP1[2].replace('.jpg','').replace("\n","")

                                # self.text_label_lp_9.setText(textLP9)
                                # self.text_label_lp_8.setText(textLP8)
                                # self.text_label_lp_7.setText(textLP7)
                                # self.text_label_lp_6.setText(textLP6)
                                # self.text_label_lp_5.setText(textLP5)
                                # self.text_label_lp_4.setText(textLP4)
                                # self.text_label_lp_3.setText(textLP3)
                                # self.text_label_lp_2.setText(textLP2)
                                # self.text_label_lp_1.setText(textLP1)
                                # self.text_label_lp_2.setText(textLP2)
                                # self.text_label_lp_3.setText(textLP3)

                                # self.text_label_lp_4.setText(textLP4)
                                # self.updateLP(textLP9=a, textLP8=a,
                                #               textLP7=a,
                                #               textLP6=a, textLP5=a,
                                #               textLP4=a,
                                #               textLP3=a, textLP2=a,
                                #               textLP1=a)


                            c = int(cls)  # integer class
                            print('x', c)
                            # TODO show text 'car 0.....'
                            # label = f'{id} {names[c]} {conf:.2f}'
                            annotator.box_label(bboxes, color=colors(c, True))

                            if c == 2 or c == 3:
                                print("BSX")

                            # if save_txt:
                            #     # to MOT format
                            #     bbox_left = output[0]
                            #     bbox_top = output[1]
                            #     bbox_w = output[2] - output[0]
                            #     bbox_h = output[3] - output[1]
                            #     # Write MOT compliant results to file
                            #     with open(txt_path, 'a') as f:
                            #         f.write(('%g ' * 10 + '\n') % (frame_idx + 1, id, bbox_left,  # MOT format
                            #                                        bbox_top, bbox_w, bbox_h, -1, -1, -1, -1))

                    for i in range(len_det):
                        # print(i)
                        MaxConf.append(det[i][4])
                        Max = max(MaxConf)
                        # maximum value
                        MaxI = MaxConf.index(Max)
                    det_max = det[[MaxI]]
                    # Write results
                    for *xyxy, conf, cls in reversed(det):
                        if save_txt:  # Write to file
                            xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(
                                -1).tolist()  # normalized xywh
                            line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                            # with open(txt_path + '.txt', 'a') as f:
                            #     f.write(('%g ' * len(line)).rstrip() % line + '\n')

                        if save_img or save_crop or view_img or self.status_license_plate_recogniton:  # Add bbox to image
                            c = int(cls)  # integer class
                            label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                            str_box = annotator.box_label(xyxy, label, color=colors(c, True))
                            t = str(xyxy)
                            print("lnn:", str(label), str(xyxy));
                            with open("temp.txt", "w") as f:
                                f.write(t + "\n")

                            # self.textBrowser_video.setText(str(count)) ###Loi in hear
                            # if self.status_license_plate_recogniton:
                            #     with open("Sample_OutPut\Image_data_management.txt", "a") as f:
                            #         f.write(file_name + "\n")
                                # self.textBrowser_video.setText(detect_lp2(xyxy, imc, file=save_dir / 'crops' / names[c] / f'{p.stem}.jpg',
                                #              BGR=True))
                            OJ =f'{names[c]}'
                            if self.status_license_plate_recogniton == True:
                                if OJ == 'car' or OJ == 'Car' or OJ == 'truck' or OJ == 'Truck' or OJ == 'motorcycle' or OJ == "Motorcycle" or OJ == 'bus' or OJ == 'Bus':
                                    save_one_box_vid(xyxy, imc, file=save_dir / 'crops' / names[
                                        c] / datetime.now().strftime(
                                        "d" + "%d%m%Y") / f'{datetime.now().strftime("%H%M%S") + "_"}.jpg',
                                                     BGR=True, Oj=OJ, LP=self.status_license_plate_recogniton)
                                elif  save_crop == True:
                                    save_one_box_vid(xyxy, imc, file=save_dir / 'crops' / names[
                                        c] / datetime.now().strftime(
                                        "d" + "%d%m%Y") / f'{datetime.now().strftime("%H%M%S") + "_"}.jpg',
                                                     BGR=True, Oj='none', LP=False)
                            if save_crop == True and self.status_license_plate_recogniton == False:
                                save_one_box_vid(xyxy, imc, file=save_dir / 'crops' / names[
                                    c] / datetime.now().strftime(
                                    "d" + "%d%m%Y") / f'{datetime.now().strftime("%H%M%S") + "_"}.jpg',
                                                 BGR=True, Oj = 'none',LP = False)
                                # print("save:",f'{p.stem}.jpg')
                                # print(xyxy)
                                # print(imc)
                # self.label_video.setText(str_box)
                # Print time (inference-only)
                LOGGER.info(f'{s}Done. ({t3 - t2:.3f}s)')
                # Stream results
                # Save results (image with detections)
                im0 = annotator.result()
                frame = im0
                resize_scale = output_size / frame.shape[0]
                frame_resized = cv2.resize(frame, (0, 0), fx=resize_scale, fy=resize_scale)

                org = (150, 150)
                font = cv2.FONT_HERSHEY_SIMPLEX
                fontScale = 3
                color = (0, 255, 0)
                thickness = 3
                frame_resized = cv2.putText(im0, str(count), org, font,
                                            fontScale, color, thickness, cv2.LINE_4)

                color = (0, 255, 0)
                start_point = (int(self.pointAx), int(self.pointAy))
                print('start_point', start_point)
                print("hn",h)
                print("wn", w)
                end_point = (int(self.pointBx), int(self.pointBy))
                print("end_point",end_point)
                # frame_resized = cv2.line(im0, start_point, end_point, color, thickness=line_thickness)
                frame_resized = cv2.rectangle(im0, start_point, end_point, color, thickness=line_thickness)

                cv2.imwrite("images/tmp/single_result_vid.jpg", frame_resized)
                # if self.checkBox_circle.checkState() > 0:
                #     # read input
                #     img = cv2.imread('images/tmp/single_org_vid.jpg')
                #
                #     # convert to gray
                #     gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                #
                #     # threshold
                #     thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)[1]
                #
                #     # find largest contour
                #     contours = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                #     contours = contours[0] if len(contours) == 2 else contours[1]
                #     if len(contours) != 0:
                #         big_contour = max(contours, key=cv2.contourArea)
                #         big_contour
                #         # fit contour to ellipse and get ellipse center, minor and major diameters and angle in degree
                #         ellipse = cv2.fitEllipse(big_contour)
                #         (xc, yc), (d1, d2), angle = ellipse
                #         # print(xc, yc, d1, d1, angle)
                #
                #         # draw ellipse
                #         result = img.copy()
                #
                #         cv2.ellipse(result, ellipse, (0, 255, 0), 3)
                #
                #         # draw circle at center
                #         xc, yc = ellipse[0]
                #         cv2.circle(result, (int(xc), int(yc)), 10, (255, 255, 255), -1)
                #         # draw vertical line
                #         # compute major radius
                #         rmajor = max(d1, d2) / 2
                #         if angle > 90:
                #             angle = angle - 90
                #         else:
                #             angle = angle + 90
                #         # print(angle)
                #         xtop = xc + math.cos(math.radians(angle)) * rmajor
                #         ytop = yc + math.sin(math.radians(angle)) * rmajor
                #         xbot = xc + math.cos(math.radians(angle + 180)) * rmajor
                #         ybot = yc + math.sin(math.radians(angle + 180)) * rmajor
                #         cv2.line(result, (int(xtop), int(ytop)), (int(xbot), int(ybot)), (0, 0, 255), 3)
                #         cv2.imwrite("images/tmp/single_result_vid.jpg", result)
                #     else:
                #         cv2.imwrite("images/tmp/single_result_vid.jpg", thresh)
                # self.label_video.setPixmap(QPixmap("test.jpg"))
                self.label_video.setPixmap(QPixmap("images/tmp/single_result_vid.jpg"))
                self.label_video.setScaledContents(True)

                # self.label_video
                # if view_img:
                # cv2.imshow(str(p), im0)
                # self.label_video.setPixmap(QPixmap("images/tmp/single_result_vid.jpg"))
                # cv2.waitKey(1)  # 1 millisecond

            if cv2.waitKey(25) & self.stopEvent.is_set() == True:
                self.stopEvent.clear()
                self.pushButton_streaming.setEnabled(True)
                self.pushButton_loadmp4.setEnabled(True)
                self.reset_vid()
                break
        # self.reset_vid()

    def updateLP(self,textLP9 = '',textLP8 = '',textLP7 = '',textLP6 = '',textLP5 = '',textLP4 = '',textLP3 = '',textLP2 = '',textLP1 = '',):
        print('uyghgggggg')
        self.text_label_lp_9.setText(textLP9)
        self.text_label_lp_8.setText(textLP8)
        self.text_label_lp_7.setText(textLP7)
        self.text_label_lp_6.setText(textLP6)
        self.text_label_lp_5.setText(textLP5)
        self.text_label_lp_4.setText(textLP4)
        self.text_label_lp_3.setText(textLP3)
        self.text_label_lp_2.setText(textLP2)
        self.text_label_lp_1.setText(textLP1)
    '''
    ### Video reset event ###
    '''

    def close_vid(self):
        self.stopEvent.set()
        self.reset_vid()
        self.buttonUpdateSettings.setEnabled(True)

    # Serial port detection
    # def port_check(self):
    #     # Detect all existing serial ports, store the information in a dictionary
    #     self.Com_Dict = {}
    #     port_list = list(serial.tools.list_ports.comports())
    #     self.box_2.clear()
    #     for port in port_list:
    #         self.Com_Dict["%s" % port[0]] = "%s" % port[1]
    #         self.box_2.addItem(port[0])
    #     if len(self.Com_Dict) == 0:
    #         self.state_label.setText(" No Serial Port")

    # Serial port information
    def port_imf(self):
        # Display details of the selected serial port
        imf_s = self.box_2.currentText()
        if imf_s != "":
            self.state_label.setText(self.Com_Dict[self.box_2.currentText()])

    # Open serial port
    def port_open(self):
        self.ser.port = self.box_2.currentText()
        self.ser.baudrate = int(self.box_3.currentText())
        self.ser.bytesize = int(self.box_4.currentText())
        self.ser.stopbits = int(self.box_6.currentText())
        self.ser.parity = self.box_5.currentText()

        try:
            self.ser.open()
        except:
            QMessageBox.critical(self, "Port Error", "This Serial Port Cannot Be Opened！")
            return None

        # Open the serial port receiving timer, the period is 2ms
        self.timer.start(2)

        if self.ser.isOpen():
            self.open_button.setEnabled(False)
            self.close_button.setEnabled(True)
            self.formGroupBox.setTitle("Serial Port Status (Opened) ")

    # Close serial port
    def port_close(self):
        self.timer.stop()
        self.timer_send.stop()
        try:
            self.ser.close()
        except:
            pass
        self.open_button.setEnabled(True)
        self.close_button.setEnabled(False)
        self.lineEdit_3.setEnabled(True)
        # The number of received data and sent data is set to zero
        self.data_num_received = 0
        # self.lineEdit.setText(str(self.data_num_received))
        self.data_num_sended = 0
        # self.lineEdit_2.setText(str(self.data_num_sended))
        self.rec_lcdNumber.display(self.data_num_received)
        self.send_lcdNumber.display(self.data_num_sended)
        self.formGroupBox.setTitle("Serial Port Status (Closed) ")

    # Send data
    def data_send(self):
        if self.ser.isOpen():
            input_s = self.send_text.toPlainText()
            if input_s != "":
                # Non-empty string
                if self.hex_send.isChecked():
                    # Hex send
                    input_s = input_s.strip()
                    send_list = []
                    while input_s != '':
                        try:
                            num = int(input_s[0:2], 16)
                        except ValueError:
                            QMessageBox.critical(self, 'wrong data',
                                                 'Please Enter Hexadecimal Data, Separated By Spaces!')
                            return None
                        input_s = input_s[2:].strip()
                        send_list.append(num)
                    input_s = bytes(send_list)
                else:
                    # Ascii send
                    input_s = (input_s + '\r\n').encode('utf-8')

                num = self.ser.write(input_s)
                self.data_num_sended += num
                # self.lineEdit_2.setText(str(self.data_num_sended))
                self.send_lcdNumber.display(self.data_num_sended)
        else:
            pass

    # Receive Data
    def data_receive(self):
        try:
            num = self.ser.inWaiting()
        except:
            self.port_close()
            return None
        if num > 0:
            data = self.ser.read(num)
            num = len(data)
            # Hex display
            if self.hex_receive.checkState():
                out_s = ''
                for i in range(0, len(data)):
                    out_s = out_s + '{:02X}'.format(data[i]) + ' '
                self.receive_text.insertPlainText(out_s)
            else:
                # The string received by the serial port is b'123', which needs to be converted into a unicode string before it can be output to the window.
                self.receive_text.insertPlainText(data.decode('iso-8859-1'))

            # Count the number of received characters
            self.data_num_received += num
            # self.lineEdit.setText(str(self.data_num_received))
            self.rec_lcdNumber.display(self.data_num_received)
            # Get the text cursor
            textCursor = self.receive_text.textCursor()
            # Scroll to bottom
            textCursor.movePosition(textCursor.End)
            # Set the cursor to the text
            self.receive_text.setTextCursor(textCursor)
        else:
            pass

    # Send data regularly
    def data_send_timer(self):
        if self.timer_send_cb.isChecked():
            self.timer_send.start(int(self.lineEdit_3.text()))
            self.lineEdit_3.setEnabled(False)
        else:
            self.timer_send.stop()
            self.lineEdit_3.setEnabled(True)

    # Clear display
    def send_data_clear(self):
        self.send_text.setText("")

    def receive_data_clear(self):
        self.receive_text.setText("")


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    myshow = MainWindows()
    myshow.show()
    sys.exit(app.exec_())
